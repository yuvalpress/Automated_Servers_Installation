#!/usr/bin/python3

# import xlrd
import sys
import subprocess
from openpyxl import load_workbook


def summon(file):
    wb = load_workbook(file, data_only=True)
    sheet = wb["Server_Configuration"]

    #set full os data
    for row in sheet.iter_rows(min_row=7, min_col=1):
        if row[1].value == 1:
            idrac_ip = row[9].value
            server_type = row[4].value
            hostname = row[6].value
            ip = row[24].value
            subnet = row[25].value
            gateway = row[26].value
            ntp = row[27].value
            dns = row[28].value
            interface = row[29].value
            interface2 = row[30].value
            network_type = row[31].value
            os = row[32].value
        else:
            continue

        if "ESXi6.5" in os:
            print("Creating customised %s ISO file for server %s." % (os, hostname))
            sub = subprocess.Popen(
                ["python3", "/mnt/c/Users/admin/Desktop/Automated "
                            "Configuration/Scripts/ISO_Scripts/ESXi_6.5_ISO_Create.py", idrac_ip, hostname, "single",
                 ip, subnet,
                 gateway, ntp, dns, interface, interface2], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            print(sub.stdout.read(), "\n")
            print("error:", sub.stderr.readlines())

        elif "RHEL7.6" in os:
            print("Creating customised %s ISO file for server %s." % (os, hostname))
            if "team" in network_type:
                popen = subprocess.Popen(
                    ["python3", "/mnt/c/Users/admin/Desktop/Automated "
                                "Configuration/Scripts/ISO_Scripts/RHEL_7.6_ISO_Create.py", idrac_ip, hostname,
                     "team", ip,
                     "%s" % subnet, gateway, ntp, dns, interface, interface2], stdout=subprocess.PIPE,
                    universal_newlines=True)
                for stdout_line in iter(popen.stdout.readline, ""):
                    print(str(stdout_line))
                popen.stdout.close()
                return_code = popen.wait()
                if return_code:
                    raise subprocess.CalledProcessError(return_code, "here")
            else:
                sub = subprocess.Popen(
                    ["python3", "/mnt/c/Users/admin/Desktop/Automated "
                                "Configuration/Scripts/ISO_Scripts/RHEL_7.6_ISO_Create.py", idrac_ip, hostname,
                     "single", ip, subnet,
                     gateway, ntp, dns, interface], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                print(sub.stdout.read(), "\n")
                print(sub.stderr.read())
        elif "RHEL7.5" in os:
            print("Creating customised %s ISO file for server %s." % (os, hostname))
            if "team" in network_type:
                sub = subprocess.Popen(
                    ["python3", "/mnt/c/Users/admin/Desktop/Automated "
                                "Configuration/Scripts/ISO_Scripts/RHEL_7.5_ISO_Create.py", idrac_ip, hostname,
                     "team", ip,
                     "%s" % subnet, gateway, ntp, dns, interface, interface2], stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE)
                print(sub.stdout.read(), "\n")
                print(sub.stderr.read())

            else:
                sub = subprocess.Popen(
                    ["python3", "/mnt/c/Users/admin/Desktop/Automated "
                                "Configuration/Scripts/ISO_Scripts/RHEL_7.5_ISO_Create.py", idrac_ip, hostname,
                     "single", ip, subnet,
                     gateway, ntp, dns, interface], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                print(sub.stdout.read(), "\n")
                print(sub.stderr.read())


if __name__ == "__main__":
    print(summon(str(sys.argv[1])))
    sys.exit()
